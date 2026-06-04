import os
import time
import shutil
from datetime import datetime
import subprocess
import pandas as pd
import requests
import base64
import hashlib
from PIL import Image
from copy import copy
from openpyxl import load_workbook
import win32com.client
import logging
from pathlib import Path
import sys


# ================== 配置区域 ==================

# 1. 根目录配置：所有脚本、Excel 文件、截图等都基于这个路径，修改这里即可适配不同环境
BASE_DIR = r"D:\刘欢_勿删\auto_report"


SCRIPT_1 = fr'"{BASE_DIR}\FileServer连通性脚本\1_FP2FileServer-REconnect.bat"'
SCRIPT_2 = fr'"{BASE_DIR}\acf_cf_cell_excel_stat\dynamic-config-file\reset_coooool.exe"'
SCRIPT_3 = fr'"{BASE_DIR}\acf_cf_cell_excel_stat\bat\startup.bat"'
SCRIPT_4 = fr'"{BASE_DIR}\temperature_excel_stat\bat\startup.bat"'

# 2. Excel 文件配置：{源文件: 目标文件}
EXCEL_PAIRS = [
    {"source": fr"{BASE_DIR}\acf_cf_cell_excel_stat\dest\acf效能统计表.xlsx", "target": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\acftest效能统计.xlsx"},
    {"source": fr"{BASE_DIR}\acf_cf_cell_excel_stat\dest\cell效能统计表.xlsx", "target": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\celltest效能统计.xlsx"},
    {"source": fr"{BASE_DIR}\acf_cf_cell_excel_stat\dest\cf效能统计表.xlsx", "target": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\cftest效能统计.xlsx"},
    {"source": fr"S:\04.FA\03.CIM\03.Schedule\AGV\ExcelUtils\file\温湿度情况统计表.xlsx", "target": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\机房温湿度统计.xlsx"},
]



# 3. 截图配置
SCREENSHOT = Path(r"D:\刘欢_勿删\auto_report\screenshots")

FILES = [
    {"file": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\acftest效能统计.xlsx", "sheet": "Sheet1", "output": fr"{SCREENSHOT}\screenshot_acftest.png", "title_rows": "29:30"},
    {"file": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\celltest效能统计.xlsx", "sheet": "Sheet1", "output": fr"{SCREENSHOT}\screenshot_celltest.png", "title_rows": "30:31"},
    {"file": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\cftest效能统计.xlsx", "sheet": "Sheet1", "output": fr"{SCREENSHOT}\screenshot_cftest.png", "title_rows": "29:30"},
    {"file": fr"{BASE_DIR}\效能表,温湿度表,fileserver表\机房温湿度统计.xlsx", "sheet": "Sheet1", "output": fr"{SCREENSHOT}\screenshot_机房温湿度.png", "title_rows": "33:35"},
]

DELAY = 1.5
# 4. 企业微信配置
WEBHOOK_URL = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=22823db5-535b-427c-b61f-502d5aa9e35c"
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}

# 5.日志配置
log_path = Path(r"D:\刘欢_勿删\auto_report\logs")
log_path.mkdir(exist_ok=True)
log_filename = log_path / f"excel_process_{datetime.now().strftime('%Y%m%d_%H%M')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
# ===============================================================


def run_scripts():
    """执行生成数据的脚本"""
    logger.info(f"正在执行脚本: {SCRIPT_1}")
    subprocess.run(f"start cmd /c call {SCRIPT_1}", shell=True)
    time.sleep(30)

    logger.info(f"正在执行脚本: {SCRIPT_2},重新获取凭证")
    subprocess.run(f"start cmd /c {SCRIPT_2}", shell=True)
    time.sleep(10)

    logger.info(f"正在执行脚本: {SCRIPT_3}")
    subprocess.run(f"start cmd /c {SCRIPT_3}", shell=True)
    logger.info("脚本执行完毕，acf、cf、cell等待文件写入磁盘...")
    time.sleep(60)

    logger.info(f"正在执行脚本：{SCRIPT_4}")
    subprocess.run(f"start cmd /c {SCRIPT_4}", shell=True)
    logger.info("脚本执行完毕，等待温湿度文件写入磁盘...")
    time.sleep(60)

    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        logger.info("正在尝试优雅关闭 Excel...")
        excel.DisplayAlerts = False
        excel.Quit()
        logger.info("Excel 已正常关闭。")
    except Exception:
        logger.info("未检测到运行的 Excel 或关闭失败，稍后将进行强制清理。")
        subprocess.run("taskkill /f /im excel.exe /t", shell=True, capture_output=True)


def process_excels():
    for pair in EXCEL_PAIRS:
        src = pair["source"]
        tgt = pair["target"]

        if not os.path.exists(src):
            logger.info(f"⚠️ 警告: 源文件不存在 -> {src}")
            continue

        try:
            df_src = pd.read_excel(src)
            if df_src.empty:
                logger.info(f"⚠️ 警告: 源文件为空 -> {src}")
                continue

            new_data = df_src.iloc[-1]
            if os.path.exists(tgt):
                wb = load_workbook(tgt)
                ws = wb["Sheet1"]
                last_row = ws.max_row
                while not ws[f"A{last_row}"].value:
                    last_row -= 1

                new_row = last_row + 1
                ws.insert_rows(new_row)
                max_col = ws.max_column
                data_len = len(new_data)
                for col in range(1, max_col + 1):
                    src_cell = ws.cell(last_row, col)
                    tgt_cell = ws.cell(new_row, col)
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.number_format = copy(src_cell.number_format)
                    tgt_cell.protection = copy(src_cell.protection)
                    tgt_cell.alignment = copy(src_cell.alignment)
                    if col <= data_len:
                        tgt_cell.value = new_data.iloc[col - 1]
                wb.save(tgt)
                logger.info(f"{tgt}写入完成")

        except Exception as e:
            logger.error(f"❌ 处理 Excel 失败: {e}")


# ====================================================================
# 导出逻辑（Excel COM 直接导出，不依赖屏幕截图，RDP 最小化也能用）

def _capture_range_as_png(ws, wb, start_row, end_row, end_col_letter, suffix, temp_dir):
    """将指定单元格区域导出为裁剪后的 PNG，返回 (路径, 宽, 高)"""
    rng = ws.Range(f"A{start_row}:{end_col_letter}{end_row}")
    rng.CopyPicture(Appearance=1, Format=2)

    # Chart 尺寸匹配 Range 实际宽高，避免画布留白造成图片间有间隙
    rng_w = int(rng.Width)
    rng_h = int(rng.Height)
    tmp_ws = wb.Worksheets.Add()
    chart_obj = tmp_ws.ChartObjects().Add(Left=0, Top=0, Width=rng_w, Height=rng_h)
    chart_obj.Chart.Paste()
    data_png = os.path.join(temp_dir, f"_{suffix}.png")
    chart_obj.Chart.Export(Filename=data_png, FilterName="PNG")
    chart_obj.Delete()
    tmp_ws.Delete()

    img = Image.open(data_png)
    gray = img.convert("L")
    bbox = gray.point(lambda x: 0 if x > 250 else 255).getbbox()
    if bbox:
        img = img.crop(bbox)
        cropped_png = os.path.join(temp_dir, f"_{suffix}_cropped.png")
        img.save(cropped_png, "PNG")
        data_png = cropped_png
    dw, dh = img.size
    img.close()
    return data_png, dw, dh


def export_excel_view(excel, output_path, delay, title_rows=None):
    """导出工作表：冻结区域的图表 + 标题行 + 最后7行数据"""
    time.sleep(delay)
    excel.ScreenUpdating = True
    time.sleep(0.5)

    temp_dir = os.path.join(os.path.dirname(output_path), "_temp_parts")
    os.makedirs(temp_dir, exist_ok=True)
    try:
        ws = excel.ActiveSheet
        wb = excel.ActiveWorkbook
        charts = ws.ChartObjects()
        parts = []

        # 1. 导出所有图表（位于冻结区域）
        for i in range(1, charts.Count + 1):
            co = charts(i)
            tp = os.path.join(temp_dir, f"_chart_{i}.png")
            co.Chart.Export(Filename=tp, FilterName="PNG")
            parts.append({
                "path": tp,
                "left": int(co.Left),
                "top": int(co.Top),
                "width": int(co.Width),
                "height": int(co.Height),
            })

        # 2. 确定列范围 & 抓取数据区域
        last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row  # xlUp
        if last_row > 0:
            data_start = max(1, last_row - 6)

            # 收集需要扫描的行范围
            scan_ranges = [range(data_start, last_row + 1)]
            title_start = title_end = None
            if title_rows:
                title_start, title_end = [int(x) for x in title_rows.split(":")]
                scan_ranges.insert(0, range(title_start, title_end + 1))

            max_col = 1
            for col in range(1, 50):
                has_data = False
                for rng in scan_ranges:
                    for row in rng:
                        v = ws.Cells(row, col).Text
                        if v and v.strip():
                            has_data = True
                            break
                    if has_data:
                        break
                if has_data:
                    max_col = col
                elif col > 10:
                    break

            end_col_letter = chr(64 + max_col) if max_col <= 26 else None
            if end_col_letter:
                data_top = max((p["top"] + p["height"] for p in parts), default=0) + 30
                data_left = min((p["left"] for p in parts), default=0)

                # 标题行（先抓，放在顶部）
                if title_start:
                    png_path, dw, dh = _capture_range_as_png(
                        ws, wb, title_start, title_end, end_col_letter, "title", temp_dir)
                    parts.append({"path": png_path, "left": data_left, "top": data_top, "width": dw, "height": dh})
                    data_top += dh

                # 最后 7 行数据
                png_path, dw, dh = _capture_range_as_png(
                    ws, wb, data_start, last_row, end_col_letter, "data", temp_dir)
                parts.append({"path": png_path, "left": data_left, "top": data_top, "width": dw, "height": dh})

        # 3. 拼接
        if not parts:
            logger.warning("没有内容可导出")
            return

        max_right = max(p["left"] + p["width"] for p in parts)
        max_bottom = max(p["top"] + p["height"] for p in parts)
        canvas = Image.new("RGB", (max_right + 20, max_bottom + 20), "white")

        for p in parts:
            img = Image.open(p["path"])
            if img.size != (p["width"], p["height"]):
                img = img.resize((p["width"], p["height"]), Image.LANCZOS)
            canvas.paste(img, (p["left"], p["top"]))
            img.close()

        canvas.save(output_path, "PNG")
        logger.info(f"导出完成: {output_path} ({canvas.size[0]}x{canvas.size[1]})")

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)



def process_one(cfg, excel):
    """处理单个文件：打开 → 导出图表+最后7行数据（RDP最小化也能用）"""
    file_path = cfg["file"]
    sheet_name = cfg["sheet"]
    output_path = cfg["output"]

    logger.info(f"\n{'='*50}")
    logger.info(f"处理: {file_path}")

    wb = excel.Workbooks.Open(file_path)
    excel.WindowState = -4137  # xlMaximized
    ws = wb.Worksheets(sheet_name)
    ws.Activate()

    # 滚到底部，确保最后几行数据在可见范围内（xlScreen模式需要）
    last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
    ws.Range(f"A{max(1, last_row - 30)}").Select()
    time.sleep(1)
    logger.info(f"滚动至 A{max(1, last_row - 30)} (最后数据行: {last_row})")

    export_excel_view(excel, output_path, DELAY, cfg.get("title_rows"))

    time.sleep(1)
    try:
        wb.Close(SaveChanges=False)
        logger.info(f"已关闭: {file_path}")
    except Exception:
        logger.info(f"关闭时出现警告（不影响结果），继续处理下一个")


def main():
    """主流程：逐个处理 FILES 列表中的每个文件"""
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        for cfg in FILES:
            try:
                process_one(cfg, excel)
            except Exception as e:
                logger.warning(f"处理失败: {cfg['file']} — {e}")
    finally:
        try:
            excel.Quit()
        except Exception:
            pass
        logger.info("\nExcel 已退出，全部截图完成")


# ===============================================================

def send_text_to_wechat(text, webhook_url):
    """发送文本消息到企业微信"""
    payload = {
        "msgtype": "text",
        "text": {
            "content": text
        }
    }
    try:
        response = requests.post(webhook_url, json=payload)
        result = response.json()
        if result.get('errcode') == 0:
            logger.info("文本消息发送成功")
        else:
            logger.error(f"文本消息发送失败: {result.get('errmsg')}")
    except Exception as e:
        logger.error(f"文本消息请求出错: {e}")


def send_image_to_wechat(image_path, webhook_url):
    """发送单张图片到企业微信"""
    if not os.path.exists(image_path):
        logger.error(f"图片文件不存在: {image_path}")
        return False

    try:
        with open(image_path, 'rb') as f:
            image_data = f.read()

        if len(image_data) > 2 * 1024 * 1024:
            logger.warning(f"图片过大 ({os.path.basename(image_path)})，跳过发送")
            return False

        base64_data = base64.b64encode(image_data).decode('utf-8')
        md5_value = hashlib.md5(image_data).hexdigest()

        payload = {
            "msgtype": "image",
            "image": {
                "base64": base64_data,
                "md5": md5_value
            }
        }

        response = requests.post(webhook_url, json=payload)
        result = response.json()

        if result.get('errcode') == 0:
            logger.info(f"推送成功: {os.path.basename(image_path)}")
            return True
        else:
            logger.error(f"推送失败: {os.path.basename(image_path)} - {result.get('errmsg')}")
            return False
    except Exception as e:
        logger.error(f"请求出错: {os.path.basename(image_path)} - {e}")
        return False


def get_all_images(directory):
    """获取指定目录下所有的图片文件路径"""
    image_list = []
    if not os.path.exists(directory):
        logger.error(f"目录不存在: {directory}")
        return image_list

    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d != "_temp_parts"]
        for file in files:
            if os.path.splitext(file)[1].lower() in IMAGE_EXTENSIONS and file != "report_file.png":
                image_list.append(os.path.join(root, file))

    return image_list


def process_shots():
    logger.info(f"开始扫描目录: {SCREENSHOT}")
    try:
    # 若没有目录则创建
        os.makedirs(SCREENSHOT, exist_ok=True)
        logger.info(f"截图保存目录已就绪: {SCREENSHOT}")
    except Exception as e:
        logger.error(f"无法创建截图目录 '{SCREENSHOT}'，原因: {e}")
        raise

    all_images = get_all_images(SCREENSHOT)

    if not all_images:
        logger.info("该目录下没有找到任何图片文件。")
        return

    logger.info(f"共找到 {len(all_images)} 张图片，开始推送...")

    success_count = 0
    for img_path in all_images:
        if send_image_to_wechat(img_path, WEBHOOK_URL):
            success_count += 1
        time.sleep(1.5)

    logger.info(f"推送任务结束！共 {len(all_images)} 张，成功 {success_count} 张。")


def get_report_time():
    """返回当天 8:00 / 12:00 / 16:00 中最接近当前时间的那个"""
    now = datetime.now()
    slots = [
        now.replace(hour=8, minute=0, second=0, microsecond=0),
        now.replace(hour=12, minute=0, second=0, microsecond=0),
        now.replace(hour=16, minute=0, second=0, microsecond=0),
    ]
    closest = min(slots, key=lambda t: abs((t - now).total_seconds()))
    return closest.strftime("%Y/%m/%d %H:%M")


if __name__ == "__main__":
    # 1. 运行脚本
    run_scripts()
    # 2. 处理数据搬运
    process_excels()
    # 3. 截图保存
    main()
    # 4. 推送图片
    report_time = get_report_time()
    report_text = f"呈长官：{report_time} ACFTEST/CFTEST/CELLTEST效能表/机房温湿度表如下"

    send_text_to_wechat(report_text, WEBHOOK_URL)
    time.sleep(1)
    process_shots()
